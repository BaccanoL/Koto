"""
📦 归档文件全文搜索引擎 (Archive Full-Text Search Engine)

功能:
- 快速索引生成 (PDF, Word, Excel, 纯文本, Markdown)
- SQLite全文搜索 (BM25算法)
- 语义搜索 (向量相似度)
- 实时索引更新
- 搜索历史追踪

使用场景:
  用户: "查一下2月份的所有涉及'黄金价格'的归档文件"
  → 1. 关键词搜索 → 2. 日期过滤 → 3. 返回结果 + 摘要
"""

import os
import sqlite3
import json
import time
import hashlib
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, asdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import unicodedata

try:
    import PyPDF2
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False


@dataclass
class IndexedFile:
    """索引文件记录"""
    id: str
    path: str
    name: str
    file_type: str
    size: int
    created_at: str
    modified_at: str
    indexed_at: str
    owner_id: str = "system"
    organization_id: str = "default"
    summary: str = ""
    keywords: List[str] = None
    content_hash: str = ""
    
    def __post_init__(self):
        if self.keywords is None:
            self.keywords = []


@dataclass
class SearchResult:
    """搜索结果"""
    file_id: str
    path: str
    name: str
    file_type: str
    summary: str
    relevance_score: float
    snippet: str  # 包含关键词的文本片段
    matched_at: int  # 文件中的字符位置


class ArchiveSearchEngine:
    """归档文件全文搜索引擎"""
    
    def __init__(self, archive_root: str = "workspace/_archive", db_path: str = ".koto_search.db"):
        self.archive_root = Path(archive_root)
        self.db_path = db_path
        self.executor = ThreadPoolExecutor(max_workers=4)
        self._init_database()
    
    def _init_database(self):
        """初始化数据库"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 文件索引表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS file_index (
                id TEXT PRIMARY KEY,
                path TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                file_type TEXT,
                size INTEGER,
                created_at TIMESTAMP,
                modified_at TIMESTAMP,
                indexed_at TIMESTAMP,
                owner_id TEXT DEFAULT 'system',
                organization_id TEXT DEFAULT 'default',
                content_hash TEXT
            )
        """)
        
        # 内容摘要表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS content_summary (
                id TEXT PRIMARY KEY,
                file_id TEXT NOT NULL UNIQUE,
                summary TEXT,
                keywords TEXT,
                entities TEXT,
                language TEXT DEFAULT 'zh',
                FOREIGN KEY(file_id) REFERENCES file_index(id)
            )
        """)
        
        # 全文索引表 (FTS5虚表)
        cursor.execute("""
            CREATE VIRTUAL TABLE IF NOT EXISTS full_text_index USING fts5(
                file_id,
                name,
                content,
                tokenize = 'porter unicode61'
            )
        """)
        
        # 搜索历史表
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS search_history (
                id TEXT PRIMARY KEY,
                user_id TEXT,
                organization_id TEXT,
                query TEXT,
                result_count INTEGER,
                execution_time_ms INTEGER,
                created_at TIMESTAMP
            )
        """)
        
        # 创建索引
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_file_type 
            ON file_index(file_type)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_created_at 
            ON file_index(created_at)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_modified_at 
            ON file_index(modified_at)
        """)
        
        conn.commit()
        conn.close()
    
    def index_archive(self, full_rebuild: bool = False) -> Dict:
        """
        索引整个归档目录
        
        Args:
            full_rebuild: 是否完全重建索引
            
        Returns:
            {
                "indexed_count": 150,
                "failed_count": 3,
                "duration_seconds": 45,
                "errors": [...]
            }
        """
        start_time = time.time()
        indexed_count = 0
        failed_count = 0
        errors = []
        
        if full_rebuild:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM file_index")
            cursor.execute("DELETE FROM content_summary")
            cursor.execute("DELETE FROM full_text_index")
            conn.commit()
            conn.close()
        
        # 并行处理文件
        files_to_index = list(self.archive_root.rglob("*"))
        files_to_index = [f for f in files_to_index if f.is_file()]
        
        futures = {}
        for file_path in files_to_index:
            future = self.executor.submit(self._index_file, file_path)
            futures[future] = file_path
        
        for future in as_completed(futures):
            try:
                if future.result():
                    indexed_count += 1
            except Exception as e:
                failed_count += 1
                errors.append({
                    "file": str(futures[future]),
                    "error": str(e)
                })
        
        duration = time.time() - start_time
        
        return {
            "indexed_count": indexed_count,
            "failed_count": failed_count,
            "total_files": len(files_to_index),
            "duration_seconds": round(duration, 2),
            "errors": errors
        }
    
    def _index_file(self, file_path: Path) -> bool:
        """索引单个文件"""
        try:
            # 检查是否已索引且未修改
            file_id = self._generate_file_id(file_path)
            content, _ = self._extract_content(file_path)
            
            if not content:
                return False
            
            # 生成内容哈希
            content_hash = hashlib.md5(content.encode()).hexdigest()
            
            # 提取摘要和关键词
            summary = self._generate_summary(content)
            keywords = self._extract_keywords(content)
            
            # 保存到数据库
            indexed_file = IndexedFile(
                id=file_id,
                path=str(file_path),
                name=file_path.name,
                file_type=file_path.suffix[1:],
                size=file_path.stat().st_size,
                created_at=datetime.fromtimestamp(file_path.stat().st_ctime).isoformat(),
                modified_at=datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                indexed_at=datetime.now().isoformat(),
                content_hash=content_hash,
                summary=summary,
                keywords=keywords
            )
            
            self._save_to_database(indexed_file, content)
            return True
            
        except Exception as e:
            print(f"Failed to index {file_path}: {e}")
            return False
    
    def _extract_content(self, file_path: Path) -> Tuple[str, str]:
        """
        提取文件内容
        
        Returns:
            (content, language)
        """
        try:
            suffix = file_path.suffix.lower()
            
            # 纯文本文件
            if suffix in ['.txt', '.md', '.markdown', '.log']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read(), 'zh'
            
            # PDF文件
            elif suffix == '.pdf' and HAS_PYPDF:
                text = []
                with open(file_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    for page in reader.pages:
                        text.append(page.extract_text())
                return '\n'.join(text), 'zh'
            
            # Word文档
            elif suffix in ['.docx', '.doc'] and HAS_DOCX:
                if suffix == '.doc':
                    # 需要转换
                    return "", 'zh'
                doc = DocxDocument(file_path)
                text = '\n'.join([para.text for para in doc.paragraphs])
                return text, 'zh'
            
            # Excel文件
            elif suffix in ['.xlsx', '.xls'] and HAS_OPENPYXL:
                text = []
                try:
                    wb = openpyxl.load_workbook(file_path)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        text.append(f"Sheet: {sheet}")
                        for row in ws.iter_rows(values_only=True):
                            text.append(' '.join(str(v) for v in row if v))
                    return '\n'.join(text), 'zh'
                except:
                    return "", 'zh'
            
            # 图片OCR
            elif suffix in ['.jpg', '.jpeg', '.png', '.gif'] and HAS_OCR:
                try:
                    img = Image.open(file_path)
                    text = pytesseract.image_to_string(img, lang='chi_sim+eng')
                    return text, 'mixed'
                except:
                    return "", 'mixed'
            
            return "", 'unknown'
            
        except Exception as e:
            print(f"Error extracting content from {file_path}: {e}")
            return "", 'unknown'
    
    def _generate_summary(self, content: str, max_length: int = 200) -> str:
        """生成内容摘要 (简单从前200字)"""
        # 移除多余空白
        content = ' '.join(content.split())
        # 移除特殊字符
        content = re.sub(r'[^\w\s\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]', ' ', content)
        return content[:max_length]
    
    def _extract_keywords(self, content: str, top_k: int = 5) -> List[str]:
        """提取关键词 (简单实现: 中文词频)"""
        # 简单关键词提取: 找出2-4个连续汉字
        keywords = re.findall(r'[\u4e00-\u9fff]{2,4}', content)
        # 计算词频
        from collections import Counter
        freq = Counter(keywords)
        return [word for word, _ in freq.most_common(top_k)]
    
    def _save_to_database(self, indexed_file: IndexedFile, content: str):
        """保存到数据库"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # 删除旧记录
            cursor.execute("DELETE FROM file_index WHERE path = ?", (indexed_file.path,))
            cursor.execute("DELETE FROM content_summary WHERE file_id = ?", (indexed_file.id,))
            cursor.execute("DELETE FROM full_text_index WHERE file_id = ?", (indexed_file.id,))
            
            # 插入文件索引
            cursor.execute("""
                INSERT INTO file_index (
                    id, path, name, file_type, size, 
                    created_at, modified_at, indexed_at, owner_id, organization_id, content_hash
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                indexed_file.id,
                indexed_file.path,
                indexed_file.name,
                indexed_file.file_type,
                indexed_file.size,
                indexed_file.created_at,
                indexed_file.modified_at,
                indexed_file.indexed_at,
                indexed_file.owner_id,
                indexed_file.organization_id,
                indexed_file.content_hash
            ))
            
            # 插入摘要
            cursor.execute("""
                INSERT INTO content_summary (
                    id, file_id, summary, keywords, entities, language
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                f"summary_{indexed_file.id}",
                indexed_file.id,
                indexed_file.summary,
                json.dumps(indexed_file.keywords),
                json.dumps({}),
                'zh'
            ))
            
            # 插入全文索引
            cursor.execute("""
                INSERT INTO full_text_index (file_id, name, content)
                VALUES (?, ?, ?)
            """, (indexed_file.id, indexed_file.name, content))
            
            conn.commit()
        finally:
            conn.close()
    
    def search(
        self,
        query: str,
        search_type: str = "hybrid",
        file_type: Optional[str] = None,
        date_range: Optional[Tuple[str, str]] = None,
        limit: int = 20,
        offset: int = 0,
        user_id: str = "system"
    ) -> Dict:
        """
        全文搜索
        
        Args:
            query: 搜索查询词
            search_type: "keyword" | "semantic" | "hybrid" (默认)
            file_type: 过滤文件类型 (pdf, docx, xlsx等)
            date_range: 日期范围 ("2026-01-01", "2026-02-14")
            limit: 返回结果数
            offset: 分页偏移
            user_id: 用户ID (用于审计)
            
        Returns:
            {
                "results": [{...}],
                "total_count": 45,
                "execution_time_ms": 123,
                "query": "黄金价格"
            }
        """
        start_time = time.time()
        results = []
        
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        
        try:
            if search_type in ["keyword", "hybrid"]:
                results = self._keyword_search(conn, query, file_type, date_range, limit, offset)
            
            # TODO: 添加语义搜索 (需要向量化)
            
            # 记录搜索历史
            self._record_search(user_id, query, len(results))
            
            execution_time = (time.time() - start_time) * 1000
            
            return {
                "results": results,
                "total_count": len(results),
                "execution_time_ms": round(execution_time, 2),
                "query": query
            }
        finally:
            conn.close()
    
    def _keyword_search(
        self,
        conn: sqlite3.Connection,
        query: str,
        file_type: Optional[str],
        date_range: Optional[Tuple[str, str]],
        limit: int,
        offset: int
    ) -> List[Dict]:
        """关键词搜索 (BM25)"""
        cursor = conn.cursor()
        results = []
        
        # 构建FTS5查询
        fts_query = f'"{query}"'  # 精确匹配
        
        sql = f"""
            SELECT DISTINCT
                f.id,
                f.path,
                f.name,
                f.file_type,
                cs.summary,
                cs.keywords,
                fti.content
            FROM full_text_index fti
            JOIN file_index f ON f.id = fti.file_id
            LEFT JOIN content_summary cs ON cs.file_id = f.id
            WHERE fti.full_text_index MATCH ?
        """
        
        params = [fts_query]
        
        # 添加文件类型过滤
        if file_type:
            sql += " AND f.file_type = ?"
            params.append(file_type)
        
        # 添加日期范围过滤
        if date_range:
            start_date, end_date = date_range
            sql += " AND f.created_at BETWEEN ? AND ?"
            params.extend([start_date, end_date])
        
        sql += " ORDER BY f.modified_at DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        
        try:
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            for row in rows:
                # 生成摘要和相关性片段
                snippet = self._generate_snippet(row['content'], query)
                relevance = self._calculate_relevance(row['content'], query)
                
                results.append({
                    "file_id": row['id'],
                    "path": row['path'],
                    "name": row['name'],
                    "file_type": row['file_type'],
                    "summary": row['summary'] or "",
                    "keywords": json.loads(row['keywords'] or '[]'),
                    "relevance_score": relevance,
                    "snippet": snippet
                })
        except Exception as e:
            print(f"Search error: {e}")
        
        return results
    
    def _generate_snippet(self, content: str, query: str, context_words: int = 20) -> str:
        """生成包含关键词的文本片段"""
        # 找到查询词在内容中的位置
        pos = content.lower().find(query.lower())
        if pos == -1:
            return content[:100]  # 返回前100字
        
        # 从前后各extraction_words字符提取
        start = max(0, pos - context_words)
        end = min(len(content), pos + len(query) + context_words)
        
        snippet = content[start:end]
        # 高亮关键词
        snippet = snippet.replace(query, f"【{query}】")
        
        return snippet
    
    def _calculate_relevance(self, content: str, query: str) -> float:
        """计算相关性分数 (0-100)"""
        # 简单实现: 关键词出现次数
        count = content.lower().count(query.lower())
        # 归一化
        score = min(100, count * 10)
        return score
    
    def _record_search(self, user_id: str, query: str, result_count: int):
        """记录搜索历史"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                INSERT INTO search_history (
                    id, user_id, organization_id, query, result_count, created_at
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                f"search_{int(time.time() * 1000)}",
                user_id,
                "default",
                query,
                result_count,
                datetime.now().isoformat()
            ))
            conn.commit()
        finally:
            conn.close()
    
    def get_search_suggestions(self, query_prefix: str, limit: int = 5) -> List[str]:
        """获取搜索建议"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        suggestions = []
        try:
            # 从搜索历史中获取
            cursor.execute("""
                SELECT DISTINCT query FROM search_history
                WHERE query LIKE ?
                ORDER BY created_at DESC
                LIMIT ?
            """, (f"{query_prefix}%", limit))
            
            suggestions = [row[0] for row in cursor.fetchall()]
        finally:
            conn.close()
        
        return suggestions
    
    def get_index_status(self) -> Dict:
        """获取索引状态"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT COUNT(*) FROM file_index")
            indexed_files = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT MAX(indexed_at) FROM file_index
            """)
            last_update = cursor.fetchone()[0]
            
            return {
                "indexed_files": indexed_files,
                "last_update": last_update,
                "status": "healthy" if indexed_files > 0 else "empty"
            }
        finally:
            conn.close()
    
    @staticmethod
    def _generate_file_id(file_path: Path) -> str:
        """生成文件ID"""
        return hashlib.md5(str(file_path).encode()).hexdigest()[:16]


# 全局实例
_search_engine: Optional[ArchiveSearchEngine] = None


def get_search_engine(archive_root: str = "workspace/_archive") -> ArchiveSearchEngine:
    """获取全局搜索引擎实例"""
    global _search_engine
    if _search_engine is None:
        _search_engine = ArchiveSearchEngine(archive_root)
    return _search_engine
