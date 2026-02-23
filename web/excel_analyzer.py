#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 数据分析器 - 专门用于处理和分析Excel文件
支持数据汇总、分组、排序、统计分析等功能
"""

import os
import sys
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import traceback


class ExcelAnalyzer:
    """Excel数据分析器"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls', '.csv']
    
    def analyze_top_customers(self, 
                             file_path: str, 
                             customer_col: str = None,
                             amount_col: str = None,
                             top_n: int = 10,
                             output_path: str = None) -> Dict[str, Any]:
        """
        分析销售数据，提取前N名客户
        
        Args:
            file_path: Excel文件路径
            customer_col: 客户名称列（自动识别如果为None）
            amount_col: 销售金额列（自动识别如果为None）
            top_n: 提取前N名（默认10）
            output_path: 输出文件路径（默认自动生成）
            
        Returns:
            {
                "success": bool,
                "result_file": str,  # 生成的结果文件路径
                "top_customers": List[Dict],  # 前N名客户数据
                "total_sales": float,  # 总销售额
                "message": str
            }
        """
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                return {"success": False, "error": f"文件不存在: {file_path}"}
            
            # 读取Excel文件
            print(f"[ExcelAnalyzer] 正在读取文件: {file_path}")
            df = pd.read_excel(file_path)
            
            if df.empty:
                return {"success": False, "error": "Excel文件为空"}
            
            print(f"[ExcelAnalyzer] 读取到 {len(df)} 行数据，列: {list(df.columns)}")
            
            # 智能识别列名
            if not customer_col:
                customer_col = self._find_column(df, ['客户', '名称', '公司', 'customer', 'name', 'company'])
            if not amount_col:
                amount_col = self._find_column(df, ['金额', '销售额', '总价', '合计', 'amount', 'sales', 'total', 'price'])
            
            # 如果没有金额列，尝试查找数量和单价列并计算
            if not amount_col:
                quantity_col = self._find_column(df, ['数量', 'quantity', 'qty', '件数'])
                price_col = self._find_column(df, ['单价', '价格', 'price', '含税单价', '不含税单价'])
                
                if quantity_col and price_col:
                    print(f"[ExcelAnalyzer] 未找到金额列，使用 {quantity_col} × {price_col} 计算金额")
                    df['_计算金额'] = pd.to_numeric(df[quantity_col], errors='coerce').fillna(0) * \
                                      pd.to_numeric(df[price_col], errors='coerce').fillna(0)
                    amount_col = '_计算金额'
            
            if not customer_col or not amount_col:
                return {
                    "success": False, 
                    "error": f"无法识别列名。可用列: {list(df.columns)}。请手动指定 customer_col 和 amount_col"
                }
            
            print(f"[ExcelAnalyzer] 使用列: 客户='{customer_col}', 金额='{amount_col}'")
            
            # 数据清洗：确保金额列是数字
            df[amount_col] = pd.to_numeric(df[amount_col], errors='coerce').fillna(0)
            
            # 按客户分组求和
            grouped = df.groupby(customer_col)[amount_col].sum().reset_index()
            grouped.columns = ['客户名称', '销售额']
            
            # 按销售额降序排序
            grouped = grouped.sort_values(by='销售额', ascending=False)
            
            # 计算总销售额
            total_sales = grouped['销售额'].sum()
            
            # 提取前N名
            top_customers = grouped.head(top_n).copy()
            
            # 计算占比
            top_customers['销售占比'] = (top_customers['销售额'] / total_sales * 100).round(2)
            top_customers['销售占比'] = top_customers['销售占比'].apply(lambda x: f"{x:.2f}%")
            
            # 生成输出文件路径
            if not output_path:
                base_name = Path(file_path).stem
                output_dir = Path(file_path).parent
                output_path = output_dir / f"{base_name}_前{top_n}客户分析.xlsx"
            
            # 保存为Excel并美化
            print(f"[ExcelAnalyzer] 正在生成结果文件: {output_path}")
            self._save_styled_excel(top_customers, output_path, f"销售额前{top_n}客户排行榜")
            
            result = {
                "success": True,
                "result_file": str(output_path),
                "top_customers": top_customers.to_dict(orient='records'),
                "total_sales": float(total_sales),
                "message": f"✅ 分析完成！前{top_n}名客户占总销售额的 {top_customers['销售额'].sum()/total_sales*100:.2f}%"
            }
            
            print(f"[ExcelAnalyzer] 分析成功: {result['message']}")
            return result
            
        except ImportError as e:
            return {
                "success": False, 
                "error": f"缺少必要的库: {str(e)}。请运行: pip install pandas openpyxl"
            }
        except Exception as e:
            traceback.print_exc()
            return {"success": False, "error": f"分析失败: {str(e)}"}
    
    def group_and_aggregate(self,
                           file_path: str,
                           group_by: str,
                           agg_col: str,
                           agg_func: str = 'sum',
                           output_path: str = None) -> Dict[str, Any]:
        """
        分组聚合分析
        
        Args:
            file_path: Excel文件路径
            group_by: 分组依据列名
            agg_col: 聚合目标列名
            agg_func: 聚合函数 (sum/mean/count/max/min)
            output_path: 输出文件路径
            
        Returns:
            分析结果字典
        """
        try:
            import pandas as pd
            
            df = pd.read_excel(file_path)
            
            # 数据清洗
            if agg_func in ['sum', 'mean', 'max', 'min']:
                df[agg_col] = pd.to_numeric(df[agg_col], errors='coerce').fillna(0)
            
            # 分组聚合
            if agg_func == 'sum':
                result_df = df.groupby(group_by)[agg_col].sum().reset_index()
                agg_name = '合计'
            elif agg_func == 'mean':
                result_df = df.groupby(group_by)[agg_col].mean().reset_index()
                agg_name = '平均'
            elif agg_func == 'count':
                result_df = df.groupby(group_by)[agg_col].count().reset_index()
                agg_name = '数量'
            elif agg_func == 'max':
                result_df = df.groupby(group_by)[agg_col].max().reset_index()
                agg_name = '最大值'
            elif agg_func == 'min':
                result_df = df.groupby(group_by)[agg_col].min().reset_index()
                agg_name = '最小值'
            else:
                return {"success": False, "error": f"不支持的聚合函数: {agg_func}"}
            
            result_df.columns = [group_by, f'{agg_col}_{agg_name}']
            result_df = result_df.sort_values(by=f'{agg_col}_{agg_name}', ascending=False)
            
            # 保存结果
            if not output_path:
                base_name = Path(file_path).stem
                output_dir = Path(file_path).parent
                output_path = output_dir / f"{base_name}_分组分析_{agg_func}.xlsx"
            
            self._save_styled_excel(result_df, output_path, f"{group_by} 分组{agg_name}分析")
            
            return {
                "success": True,
                "result_file": str(output_path),
                "data": result_df.to_dict(orient='records'),
                "message": f"✅ 分组分析完成，共 {len(result_df)} 个分组"
            }
            
        except Exception as e:
            traceback.print_exc()
            return {"success": False, "error": f"分组分析失败: {str(e)}"}
    
    def calculate_statistics(self, file_path: str, columns: List[str] = None) -> Dict[str, Any]:
        """
        计算统计信息
        
        Args:
            file_path: Excel文件路径
            columns: 要分析的列名列表（None表示所有数值列）
            
        Returns:
            统计结果字典
        """
        try:
            import pandas as pd
            
            df = pd.read_excel(file_path)
            
            if columns:
                df = df[columns]
            else:
                # 只选择数值列
                df = df.select_dtypes(include=['number'])
            
            if df.empty:
                return {"success": False, "error": "没有可分析的数值列"}
            
            # 计算统计信息
            stats = df.describe().T
            stats['总和'] = df.sum()
            stats['中位数'] = df.median()
            
            return {
                "success": True,
                "statistics": stats.to_dict(),
                "message": f"✅ 已计算 {len(df.columns)} 列的统计信息"
            }
            
        except Exception as e:
            return {"success": False, "error": f"统计计算失败: {str(e)}"}
    
    def smart_analyze(self, file_path: str, question: str) -> Dict[str, Any]:
        """
        智能分析：根据问题自动选择分析方法
        
        Args:
            file_path: Excel文件路径
            question: 分析需求描述
            
        Returns:
            分析结果
        """
        question_lower = question.lower()
        
        # 判断分析类型
        if any(kw in question_lower for kw in ['前', '前十', 'top', '排名', '排行']):
            # 提取数量
            import re
            match = re.search(r'前(\d+)', question)
            top_n = int(match.group(1)) if match else 10
            return self.analyze_top_customers(file_path, top_n=top_n)
        
        elif any(kw in question_lower for kw in ['分组', '按', '统计', 'group']):
            # 需要更多参数，返回提示
            return {
                "success": False,
                "error": "分组分析需要指定 group_by 和 agg_col 参数",
                "hint": "请使用 group_and_aggregate 方法"
            }
        
        elif any(kw in question_lower for kw in ['统计', '平均', '总和', 'statistics']):
            return self.calculate_statistics(file_path)
        
        else:
            # 默认：提取前10客户
            return self.analyze_top_customers(file_path, top_n=10)
    
    # ======== 辅助方法 ========
    
    def _find_column(self, df, keywords: List[str]) -> Optional[str]:
        """智能匹配列名"""
        cols = df.columns.astype(str)
        for col in cols:
            col_lower = col.lower()
            for keyword in keywords:
                if keyword.lower() in col_lower:
                    return col
        # 如果没找到，尝试模糊匹配
        for col in cols:
            for keyword in keywords:
                if keyword.lower()[:2] in col.lower():
                    return col
        return None
    
    def _save_styled_excel(self, df, output_path: str, sheet_name: str = "Sheet1"):
        """保存带样式的Excel文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 标题行样式
                    if r_idx == 1:
                        cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.font = Font(name='微软雅黑', size=11)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        # 数值列右对齐
                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 添加边框
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
            
            # 调整列宽
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存文件
            wb.save(output_path)
            print(f"[ExcelAnalyzer] Excel文件已保存: {output_path}")
            
        except Exception as e:
            print(f"[ExcelAnalyzer] 保存Excel失败: {e}")
            # 降级：使用pandas保存
            df.to_excel(output_path, index=False)


# 便捷函数
def analyze_excel(file_path: str, analysis_type: str = "top_customers", **kwargs) -> Dict[str, Any]:
    """
    便捷函数：分析Excel文件
    
    Args:
        file_path: Excel文件路径
        analysis_type: 分析类型 (top_customers/group_aggregate/statistics/smart)
        **kwargs: 其他参数传递给具体方法
        
    Returns:
        分析结果字典
    """
    analyzer = ExcelAnalyzer()
    
    if analysis_type == "top_customers":
        return analyzer.analyze_top_customers(file_path, **kwargs)
    elif analysis_type == "group_aggregate":
        return analyzer.group_and_aggregate(file_path, **kwargs)
    elif analysis_type == "statistics":
        return analyzer.calculate_statistics(file_path, **kwargs)
    elif analysis_type == "smart":
        question = kwargs.get('question', '')
        return analyzer.smart_analyze(file_path, question)
    else:
        return {"success": False, "error": f"未知的分析类型: {analysis_type}"}


# 测试代码
if __name__ == "__main__":
    print("Excel分析器测试")
    print("=" * 60)
    
    # 测试文件路径
    test_file = r"workspace\销售台账.xlsx"
    
    if os.path.exists(test_file):
        analyzer = ExcelAnalyzer()
        result = analyzer.analyze_top_customers(test_file, top_n=10)
        
        if result['success']:
            print(f"✅ {result['message']}")
            print(f"📊 总销售额: {result['total_sales']:,.2f}")
            print(f"📁 结果文件: {result['result_file']}")
            print("\n前10名客户:")
            for i, customer in enumerate(result['top_customers'], 1):
                print(f"{i}. {customer['客户名称']}: {customer['销售额']:,.2f} ({customer['销售占比']})")
        else:
            print(f"❌ 分析失败: {result.get('error', '未知错误')}")
    else:
        print(f"测试文件不存在: {test_file}")
