"""
Batch file operations with progress tracking.
"""
import os
import re
import json
import csv
import uuid
import shutil
import queue
import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


@dataclass
class BatchJobRecord:
    job_id: str
    name: str
    operation: str
    input_dir: str
    output_dir: str
    options: Dict[str, Any] = field(default_factory=dict)
    status: str = "queued"
    total_items: int = 0
    processed_items: int = 0
    failed_items: int = 0
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    results: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


class BatchFileOpsManager:
    def __init__(self):
        self.jobs: Dict[str, BatchJobRecord] = {}
        self.job_events: Dict[str, queue.Queue] = {}
        self.lock = threading.Lock()

    def list_jobs(self) -> List[Dict[str, Any]]:
        with self.lock:
            return [self._job_to_dict(job) for job in self.jobs.values()]

    def get_job(self, job_id: str) -> Optional[Dict[str, Any]]:
        with self.lock:
            job = self.jobs.get(job_id)
            return self._job_to_dict(job) if job else None

    def create_job(self, name: str, operation: str, input_dir: str, output_dir: str, options: Dict[str, Any]) -> BatchJobRecord:
        job_id = uuid.uuid4().hex[:12]
        job = BatchJobRecord(
            job_id=job_id,
            name=name,
            operation=operation,
            input_dir=input_dir,
            output_dir=output_dir,
            options=options or {},
        )
        with self.lock:
            self.jobs[job_id] = job
            self.job_events[job_id] = queue.Queue()
        return job

    def start_job(self, job_id: str):
        thread = threading.Thread(target=self._run_job, args=(job_id,), daemon=True)
        thread.start()

    def stream_job(self, job_id: str) -> Iterable[str]:
        for event in self.iter_job_events(job_id):
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

    def iter_job_events(self, job_id: str) -> Iterable[Dict[str, Any]]:
        q = self.job_events.get(job_id)
        if not q:
            yield {"type": "error", "message": "job not found"}
            return
        while True:
            event = q.get()
            yield event
            if event.get("type") == "final":
                break

    def parse_command(self, user_input: str) -> Dict[str, Any]:
        text = (user_input or "").strip()
        op = self._detect_operation(text)
        if not op:
            return {
                "success": False,
                "error": "æœªè¯†åˆ«æ‰¹é‡æ“ä½œç±»åž‹",
                "hint": self._usage_hint()
            }

        input_dir = self._extract_path(text, ["è¾“å…¥", "ä»Ž", "ç›®å½•", "æ–‡ä»¶å¤¹"])
        output_dir = self._extract_path(text, ["è¾“å‡º", "ä¿å­˜åˆ°", "åˆ°", "ç›®æ ‡", "å¯¼å‡ºåˆ°"])
        if not input_dir or not output_dir:
            return {
                "success": False,
                "error": "ç¼ºå°‘è¾“å…¥æˆ–è¾“å‡ºç›®å½•",
                "hint": self._usage_hint()
            }

        options: Dict[str, Any] = {}
        options["include_exts"] = self._extract_exts(text)

        if op == "convert":
            target_ext = self._extract_target_ext(text)
            if not target_ext:
                return {
                    "success": False,
                    "error": "ç¼ºå°‘ç›®æ ‡æ ¼å¼",
                    "hint": "ç¤ºä¾‹: æ‰¹é‡è½¬æ¢ ä»Ž C:\\A åˆ° D:\\B è½¬ä¸º pdf"
                }
            options["target_ext"] = target_ext

        if op == "rename":
            options.update(self._extract_rename_rules(text))

        if op == "compress_images":
            options.update(self._extract_image_options(text))

        if op == "extract_text":
            options["target_ext"] = ".txt"

        return {
            "success": True,
            "operation": op,
            "input_dir": input_dir,
            "output_dir": output_dir,
            "options": options,
        }

    def is_batch_command(self, user_input: str) -> bool:
        text = (user_input or "").strip()
        return self._detect_operation(text) is not None

    def _detect_operation(self, text: str) -> Optional[str]:
        if any(k in text for k in ["æ‰¹é‡è½¬æ¢", "æ ¼å¼è½¬æ¢", "æ‰¹é‡è½¬", "è½¬æˆ", "è½¬æ¢ä¸º"]):
            return "convert"
        if any(k in text for k in ["æ‰¹é‡é‡å‘½å", "é‡å‘½å", "æ”¹å", "æ‰¹é‡æ”¹å"]):
            return "rename"
        if any(k in text for k in ["æ‰¹é‡å½’æ¡£", "æ‰¹é‡æ•´ç†", "å½’æ¡£", "æ•´ç†æ–‡ä»¶"]):
            return "organize"
        if any(k in text for k in ["æ‰¹é‡åŽ‹ç¼©", "åŽ‹ç¼©å›¾ç‰‡", "å›¾ç‰‡åŽ‹ç¼©", "åŽ‹ç¼©å›¾ç‰‡"]):
            return "compress_images"
        if any(k in text for k in ["æ‰¹é‡æå–", "æŠ½å–æ–‡æœ¬", "æå–æ–‡æœ¬", "OCR"]):
            return "extract_text"
        if any(k in text for k in ["æ‰¹é‡æ¸…ç†", "è§„èŒƒåŒ–", "æ¸…ç†", "æ ‡å‡†åŒ–"]):
            return "clean_normalize"
        return None

    def _extract_path(self, text: str, anchors: List[str]) -> Optional[str]:
        quoted = re.findall(r"[\"']([^\"']+)[\"']", text)
        if quoted:
            # Prefer the first quoted path
            return quoted[0]

        for anchor in anchors:
            pattern = rf"{anchor}\s*[:ï¼š]?\s*([^\s]+)"
            match = re.search(pattern, text)
            if match:
                return match.group(1)

        # Fallback to Windows absolute path
        match = re.search(r"([A-Za-z]:\\[^\s]+)", text)
        if match:
            return match.group(1)
        return None

    def _extract_exts(self, text: str) -> List[str]:
        exts = re.findall(r"\.(docx|pdf|xlsx|pptx|txt|json|csv|png|jpg|jpeg)", text, re.IGNORECASE)
        if not exts:
            exts = re.findall(r"(docx|pdf|xlsx|pptx|txt|json|csv|png|jpg|jpeg)", text, re.IGNORECASE)
        return [f".{e.lower()}" for e in exts]

    def _extract_target_ext(self, text: str) -> Optional[str]:
        match = re.search(r"è½¬ä¸º\s*([A-Za-z0-9.]+)", text)
        if not match:
            match = re.search(r"è½¬æ¢ä¸º\s*([A-Za-z0-9.]+)", text)
        if not match:
            match = re.search(r"è½¬æˆ\s*([A-Za-z0-9.]+)", text)
        if not match:
            return None
        value = match.group(1).strip().lower()
        if not value.startswith("."):
            value = f".{value}"
        return value

    def _extract_rename_rules(self, text: str) -> Dict[str, Any]:
        rules: Dict[str, Any] = {}
        prefix = self._extract_kv(text, "å‰ç¼€")
        suffix = self._extract_kv(text, "åŽç¼€")
        seq = self._extract_kv(text, "åºå·")
        replace = self._extract_kv(text, "æ›¿æ¢")

        if prefix:
            rules["prefix"] = prefix
        if suffix:
            rules["suffix"] = suffix
        if seq:
            rules["seq_start"] = seq
        if replace and "->" in replace:
            old, new = replace.split("->", 1)
            rules["replace"] = (old, new)
        return rules

    def _extract_kv(self, text: str, key: str) -> Optional[str]:
        match = re.search(rf"{key}\s*[=:ï¼š]\s*([^\s]+)", text)
        if match:
            return match.group(1)
        return None

    def _extract_image_options(self, text: str) -> Dict[str, Any]:
        options: Dict[str, Any] = {}
        quality = self._extract_kv(text, "è´¨é‡")
        width = self._extract_kv(text, "å®½")
        height = self._extract_kv(text, "é«˜")
        fmt = self._extract_kv(text, "æ ¼å¼")
        if quality and quality.isdigit():
            options["quality"] = int(quality)
        if width and width.isdigit():
            options["width"] = int(width)
        if height and height.isdigit():
            options["height"] = int(height)
        if fmt:
            fmt = fmt.lower()
            if not fmt.startswith("."):
                fmt = f".{fmt}"
            options["format"] = fmt
        return options

    def _usage_hint(self) -> str:
        return (
            "ç¤ºä¾‹:\n"
            "- æ‰¹é‡è½¬æ¢ ä»Ž \"C:\\A\" è¾“å‡ºåˆ° \"D:\\B\" è½¬ä¸º pdf\n"
            "- æ‰¹é‡é‡å‘½å ä»Ž \"C:\\A\" è¾“å‡ºåˆ° \"D:\\B\" å‰ç¼€=é¡¹ç›®_ åºå·=001\n"
            "- æ‰¹é‡å½’æ¡£ ä»Ž \"C:\\A\" è¾“å‡ºåˆ° \"D:\\B\"\n"
            "- å›¾ç‰‡åŽ‹ç¼© ä»Ž \"C:\\A\" è¾“å‡ºåˆ° \"D:\\B\" è´¨é‡=80 å®½=1200\n"
            "- æŠ½å–æ–‡æœ¬ ä»Ž \"C:\\A\" è¾“å‡ºåˆ° \"D:\\B\"\n"
            "- æ‰¹é‡æ¸…ç† ä»Ž \"C:\\A\" è¾“å‡ºåˆ° \"D:\\B\""
        )

    def _run_job(self, job_id: str):
        with self.lock:
            job = self.jobs.get(job_id)
        if not job:
            return

        job.status = "running"
        job.started_at = datetime.now().isoformat()
        self._emit(job_id, {
            "type": "progress",
            "current": 0,
            "total": 0,
            "status": "start",
            "detail": f"å¼€å§‹å¤„ç†: {job.name}"
        })

        input_dir = Path(job.input_dir)
        output_dir = Path(job.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        files = self._collect_files(input_dir, job.options.get("include_exts"))
        job.total_items = len(files)

        self._emit(job_id, {
            "type": "progress",
            "current": 0,
            "total": job.total_items,
            "status": "scan",
            "detail": f"å‘çŽ° {job.total_items} ä¸ªæ–‡ä»¶"
        })

        for idx, path in enumerate(files, start=1):
            try:
                result = self._process_file(job, path, input_dir, output_dir, idx)
                job.results.append(result)
                if result.get("success"):
                    job.processed_items += 1
                else:
                    job.failed_items += 1
            except Exception as e:
                job.failed_items += 1
                job.errors.append(str(e))
                result = {"source": str(path), "success": False, "error": str(e)}
                job.results.append(result)

            self._emit(job_id, {
                "type": "progress",
                "current": idx,
                "total": job.total_items,
                "status": "processing",
                "detail": f"{path.name}"
            })

        job.status = "completed" if job.failed_items == 0 else "failed"
        job.completed_at = datetime.now().isoformat()

        summary = self._build_summary(job)
        self._emit(job_id, {
            "type": "final",
            "summary": summary,
            "job": self._job_to_dict(job)
        })

    def _emit(self, job_id: str, event: Dict[str, Any]):
        q = self.job_events.get(job_id)
        if q:
            q.put(event)

    def _collect_files(self, input_dir: Path, include_exts: Optional[List[str]]) -> List[Path]:
        if not input_dir.exists():
            return []
        files = [p for p in input_dir.rglob('*') if p.is_file()]
        if include_exts:
            include_exts = [e.lower() for e in include_exts]
            files = [p for p in files if p.suffix.lower() in include_exts]
        return files

    def _process_file(self, job: BatchJobRecord, path: Path, input_root: Path, output_root: Path, index: int) -> Dict[str, Any]:
        op = job.operation
        if op == "convert":
            return self._convert_file(path, input_root, output_root, job.options)
        if op == "rename":
            return self._rename_file(path, input_root, output_root, job.options, index)
        if op == "organize":
            return self._organize_file(path, output_root)
        if op == "compress_images":
            return self._compress_image(path, input_root, output_root, job.options)
        if op == "extract_text":
            return self._extract_text(path, input_root, output_root)
        if op == "clean_normalize":
            return self._clean_normalize(path, input_root, output_root)
        return {"source": str(path), "success": False, "error": f"æœªçŸ¥æ“ä½œ: {op}"}

    def _relative_output_path(self, path: Path, input_root: Path, output_root: Path, new_ext: Optional[str] = None) -> Path:
        relative = path.relative_to(input_root)
        if new_ext:
            relative = relative.with_suffix(new_ext)
        dest = output_root / relative
        dest.parent.mkdir(parents=True, exist_ok=True)
        return dest

    def _convert_file(self, path: Path, input_root: Path, output_root: Path, options: Dict[str, Any]) -> Dict[str, Any]:
        target_ext = options.get("target_ext")
        if not target_ext:
            return {"source": str(path), "success": False, "error": "ç¼ºå°‘ç›®æ ‡æ ¼å¼"}

        source_ext = path.suffix.lower()
        dest = self._relative_output_path(path, input_root, output_root, target_ext)

        try:
            if source_ext in [".docx", ".doc"] and target_ext == ".txt":
                from docx import Document
                doc = Document(str(path))
                text = "\n".join([p.text for p in doc.paragraphs])
                dest.write_text(text, encoding="utf-8")
                return {"source": str(path), "success": True, "output": str(dest)}

            if source_ext == ".pdf" and target_ext == ".txt":
                import PyPDF2
                text = ""
                with open(path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    for page in reader.pages:
                        text += page.extract_text() or ""
                dest.write_text(text, encoding="utf-8")
                return {"source": str(path), "success": True, "output": str(dest)}

            if source_ext in [".xlsx", ".xls"] and target_ext == ".csv":
                import openpyxl
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                with open(dest, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(list(row))
                return {"source": str(path), "success": True, "output": str(dest)}

            if source_ext == ".csv" and target_ext in [".xlsx", ".xls"]:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        ws.append(row)
                wb.save(dest)
                return {"source": str(path), "success": True, "output": str(dest)}

            if source_ext in [".png", ".jpg", ".jpeg"] and target_ext in [".png", ".jpg", ".jpeg"]:
                from PIL import Image
                image = Image.open(path)
                if target_ext in [".jpg", ".jpeg"] and image.mode != "RGB":
                    image = image.convert("RGB")
                image.save(dest)
                return {"source": str(path), "success": True, "output": str(dest)}

            if source_ext in [".docx", ".doc"] and target_ext == ".pdf":
                try:
                    from docx2pdf import convert
                    dest.parent.mkdir(parents=True, exist_ok=True)
                    convert(str(path), str(dest))
                    return {"source": str(path), "success": True, "output": str(dest)}
                except Exception as e:
                    return {"source": str(path), "success": False, "error": f"docx->pdf éœ€è¦ docx2pdf: {e}"}

            return {"source": str(path), "success": False, "error": f"ä¸æ”¯æŒ {source_ext} -> {target_ext}"}
        except Exception as e:
            return {"source": str(path), "success": False, "error": str(e)}

    def _rename_file(self, path: Path, input_root: Path, output_root: Path, options: Dict[str, Any], index: int) -> Dict[str, Any]:
        prefix = options.get("prefix", "")
        suffix = options.get("suffix", "")
        seq_start = options.get("seq_start")
        replace = options.get("replace")

        stem = path.stem
        if replace and isinstance(replace, Tuple):
            stem = stem.replace(replace[0], replace[1])

        seq_part = ""
        if seq_start and seq_start.isdigit():
            seq_width = len(seq_start)
            seq_part = str(int(seq_start) + index - 1).zfill(seq_width)

        new_name = f"{prefix}{stem}{suffix}{seq_part}{path.suffix}"
        dest = self._relative_output_path(path, input_root, output_root)
        dest = dest.with_name(new_name)
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, dest)
        return {"source": str(path), "success": True, "output": str(dest)}

    def _organize_file(self, path: Path, output_root: Path) -> Dict[str, Any]:
        try:
            from web.file_analyzer import FileAnalyzer
            from web.file_organizer import FileOrganizer
        except ImportError:
            from file_analyzer import FileAnalyzer
            from file_organizer import FileOrganizer

        analyzer = FileAnalyzer()
        analysis = analyzer.analyze_file(str(path))
        folder = analysis.get("suggested_folder") or "other"
        organizer = FileOrganizer(str(output_root))
        result = organizer.organize_file(str(path), folder, auto_confirm=True, metadata={
            "entity": analysis.get("entity"),
            "entity_type": analysis.get("entity_type")
        })
        if result.get("success"):
            return {"source": str(path), "success": True, "output": result.get("dest_file")}
        return {"source": str(path), "success": False, "error": result.get("error", "å½’æ¡£å¤±è´¥")}

    def _compress_image(self, path: Path, input_root: Path, output_root: Path, options: Dict[str, Any]) -> Dict[str, Any]:
        if path.suffix.lower() not in [".png", ".jpg", ".jpeg"]:
            return {"source": str(path), "success": False, "error": "éžå›¾ç‰‡æ–‡ä»¶"}
        try:
            from PIL import Image
        except Exception as e:
            return {"source": str(path), "success": False, "error": f"ç¼ºå°‘ Pillow: {e}"}

        quality = options.get("quality", 80)
        width = options.get("width")
        height = options.get("height")
        target_fmt = options.get("format") or path.suffix.lower()
        dest = self._relative_output_path(path, input_root, output_root, target_fmt)

        image = Image.open(path)
        if width or height:
            new_w = width or image.size[0]
            new_h = height or image.size[1]
            image = image.resize((new_w, new_h))
        if target_fmt in [".jpg", ".jpeg"] and image.mode != "RGB":
            image = image.convert("RGB")
        image.save(dest, quality=quality, optimize=True)
        return {"source": str(path), "success": True, "output": str(dest)}

    def _extract_text(self, path: Path, input_root: Path, output_root: Path) -> Dict[str, Any]:
        dest = self._relative_output_path(path, input_root, output_root, ".txt")
        try:
            from web.file_analyzer import FileAnalyzer
        except ImportError:
            from file_analyzer import FileAnalyzer

        analyzer = FileAnalyzer()
        text = analyzer._extract_content(str(path))
        if not text:
            return {"source": str(path), "success": False, "error": "æ— æ³•æå–æ–‡æœ¬"}
        dest.write_text(text, encoding="utf-8")
        return {"source": str(path), "success": True, "output": str(dest)}

    def _clean_normalize(self, path: Path, input_root: Path, output_root: Path) -> Dict[str, Any]:
        ext = path.suffix.lower()
        dest = self._relative_output_path(path, input_root, output_root)

        if ext in [".txt", ".md", ".csv", ".json", ".log"]:
            content = path.read_text(encoding="utf-8", errors="ignore")
            content = "\n".join(line.rstrip() for line in content.splitlines())
            dest.write_text(content, encoding="utf-8")
            return {"source": str(path), "success": True, "output": str(dest)}

        if ext in [".png", ".jpg", ".jpeg"]:
            try:
                from PIL import Image
            except Exception as e:
                return {"source": str(path), "success": False, "error": f"ç¼ºå°‘ Pillow: {e}"}
            image = Image.open(path)
            if image.mode != "RGB" and ext in [".jpg", ".jpeg"]:
                image = image.convert("RGB")
            image.save(dest)
            return {"source": str(path), "success": True, "output": str(dest)}

        shutil.copy2(path, dest)
        return {"source": str(path), "success": True, "output": str(dest)}

    def _build_summary(self, job: BatchJobRecord) -> str:
        lines = [
            f"ðŸ“¦ æ‰¹é‡ä»»åŠ¡å®Œæˆ: {job.name}",
            f"- çŠ¶æ€: {job.status}",
            f"- æ€»æ•°: {job.total_items}",
            f"- æˆåŠŸ: {job.processed_items}",
            f"- å¤±è´¥: {job.failed_items}",
            f"- è¾“å‡ºç›®å½•: {job.output_dir}",
        ]
        if job.failed_items > 0:
            lines.append("\nâ— å¤±è´¥æ ·ä¾‹:")
            for item in job.results:
                if not item.get("success"):
                    lines.append(f"- {Path(item.get('source', '')).name}: {item.get('error', 'error')}")
                    if len(lines) > 12:
                        break
        return "\n".join(lines)

    def _job_to_dict(self, job: Optional[BatchJobRecord]) -> Optional[Dict[str, Any]]:
        if not job:
            return None
        return {
            "job_id": job.job_id,
            "name": job.name,
            "operation": job.operation,
            "input_dir": job.input_dir,
            "output_dir": job.output_dir,
            "status": job.status,
            "total_items": job.total_items,
            "processed_items": job.processed_items,
            "failed_items": job.failed_items,
            "created_at": job.created_at,
            "started_at": job.started_at,
            "completed_at": job.completed_at,
        }
